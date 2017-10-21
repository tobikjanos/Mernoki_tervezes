import psycopg2
import sys, re
from openpyxl import Workbook

# init worksheet
wb = Workbook()
ws = wb.active

#columnA = ws['A']	# A is 0
#columnB = ws[1]	# B is 1
#columnC = ws[2]	# C is 2
#columnD = ws[3]	# D is 3

# init database connection
conn = psycopg2.connect( database="lavinia", user="postgres", password="qaswed123", host="127.0.0.1", port="5432" )

cur = conn.cursor()

query = "select substring(source_link_no from 11 for 5) as ndb_no, count( substring(source_link_no from 11 for 5) ) \
from minta2.food_source                                          		\
where source_link_no != '' and source_link_no not like 'USDA:0 %'	\
group by substring(source_link_no from 11 for 5)	\
having count( substring(source_link_no from 11 for 5) ) > 1	\
order by substring(source_link_no from 11 for 5)"

ws['A1'] = "NDB_No"
ws['B1'] = "Food_ID"
ws['C1'] = "Magyar név"
ws['D1'] = "Angol név"



rowCntr = 2

#execute query
cur.execute(query)
#fetch rows into list
list = cur.fetchall()
#display results
for row in list:
	ndb_no = row[0]
	
	# insert ndb_no to worksheet
	ws['A' + str(rowCntr)] = ndb_no
	
	query = "select food_id from minta2.food_source where source_link_no like 'USDA%' || '" + ndb_no + "' || ' SOTE%'"
	#execute query
	cur.execute(query)
	#fetch rows into list
	listFoodID = cur.fetchall()
	#display results
	for idx in listFoodID:
		
		food_id = idx[0]
		
		# insert food_id to worksheet
		ws['B' + str(rowCntr)] = food_id
		
		query = "select _f.food_id, _lt.label_text	\
					   from minta2.label_text _lt inner join minta2.label _l on _lt.label_id = _l.label_id			\
					   inner join minta2.food _f on _l.label_id = _f.foodname_label_id									\
					   where _f.food_id = " + str(food_id) + " and _lt.lang_id = 1"
		
		#execute query
		cur.execute(query)
		#fetch rows into list
		food_HU = cur.fetchone()
		
		# insert label of food to worksheet, lang: HU
		if not food_HU == None:
			ws['C' + str(rowCntr)] = food_HU[1]
		
		query = "select _f.food_id, _lt.label_text	\
					   from minta2.label_text _lt inner join minta2.label _l on _lt.label_id = _l.label_id			\
					   inner join minta2.food _f on _l.label_id = _f.foodname_label_id									\
					   where _f.food_id = " + str(food_id) + " and _lt.lang_id = 2"
		
		#execute query
		cur.execute(query)
		#fetch rows into list
		food_EN = cur.fetchone()
		
		# insert label of food to worksheet, lang: EN
		if not food_EN == None:
			ws['D' + str(rowCntr)] = food_EN[1]
		
		# increment counter
		rowCntr = rowCntr + 1
		
		#save workbook
		wb.save("lista.xlsx")
			
	

conn.commit()
conn.close()